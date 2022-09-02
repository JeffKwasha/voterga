# handle importing precinct data from the sos website
import logging
from pathlib import Path
from db import Name

# https://xlrd.readthedocs.io/en/latest/ (old xls)
# https://openpyxl.readthedocs.io/en/stable/ (new xlsx)
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException


class Xlsx:
    """ Handle loading xls and generate objects
    """
    def __init__(self, filename: Path, read_only: bool = False):
        try:
            self.wb: Workbook = load_workbook(filename=filename, read_only=read_only)
            self._filename = filename
            self._max_column = None
            self._row_names = []
        except InvalidFileException as e:
            logging.error(f"Unable to open {filename}: {e}")

    @property
    def max_column(self) -> int:
        """ return the name of valid columns """
        if self._max_column is not None:
            return self._max_column
        ws = self.wb.active
        stop = False
        for n, c in enumerate(next(ws.iter_rows())):
            if c and c.value:
                self._max_column, stop = n, False
                continue
            elif stop:
                break
            stop = True
        return self._max_column

    @property
    def row_names(self) -> list:
        """ return a list of values from the first column """
        if self._row_names:
            return self._row_names
        ws = self.wb.active
        stop = False
        for n, v in enumerate(next(ws.iter_cols())):
            if not v:
                logging.error(f"cell is invalid! {n}:{v}")
                continue
            row, col, value = v.row, v.column, v.value
            self._row_names.append(value)
            if value:
                stop = False
                continue
            elif stop and any(self._row_names):
                break
            stop = True
        return self._row_names

    def load_columns(self, obj: callable, *args, **kwargs) -> list:
        """ Generate an object for each column where each row is a (potential) parameter:
        name    foo     bar
        color   red     black
        flavor  cherry  raspberry
        ---
        obj(**{name: foo, color: red, flavor: cherry}, _file=FILENAME, _column=int(column)),
        obj(**{name: bar, color: black, flavor: raspberry}, _file=FILENAME, _column=int(column)),
        """

        # load the 1st column using names (or Name)
        ws: Worksheet = self.wb.active

        def build_object(col: int, obj: callable):
            vals = {'_file': self._filename, '_column': col}
            for n, v in enumerate(self.row_names):
                if not v:
                    continue
                vals[f"{n}:{v}"] = ws.cell(row=n+1, column=col).value
            return obj(*args, **kwargs, **vals)

        return [build_object(col, obj=obj) for col in range(2, self.max_column)]

